from openpyxl import load_workbook
import pandas as pd  # pandas is used for NaN checking

def to_int(value):
    """Attempt to convert a value to int, handling NaN and non-numeric values."""
    try:
        if pd.isna(value):
            return None
        return int(value)
    except (ValueError, TypeError):
        return None

def normalize_column(ws, column_name):
    """Convert all values in the specified column to integers."""
    column_index = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == column_name:
            column_index = col[0].column
            break
    
    if column_index is None:
        print(f"Column '{column_name}' not found.")
        return
    
    for row in ws.iter_rows(min_row=2, max_col=column_index, max_row=ws.max_row):
        cell = row[column_index - 1]  # Adjust for zero-based indexing
        cell.value = to_int(cell.value)

def find_and_delete_duplicates(file_path, phone_column_name, mobile_column_name):
    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Normalize the Phone and Mobile columns
    normalize_column(ws, phone_column_name)
    normalize_column(ws, mobile_column_name)
    
    # Find the column indices for the specified column names
    phone_column_index = None
    mobile_column_index = None
    name_column_index = None
    
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == phone_column_name:
            phone_column_index = col[0].column
        elif col[0].value == mobile_column_name:
            mobile_column_index = col[0].column
        elif col[0].value == 'Name':
            name_column_index = col[0].column
    
    if phone_column_index is None or mobile_column_index is None or name_column_index is None:
        print("One or more columns not found.")
        return

    seen = {}
    duplicates_info = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        phone_cell = row[phone_column_index - 1]  # Adjust for zero-based indexing
        mobile_cell = row[mobile_column_index - 1]
        name_cell = row[name_column_index - 1]  # Get the name cell
        phone_value_int = to_int(phone_cell.value)  # Convert phone value to int if possible
        mobile_value_int = to_int(mobile_cell.value)  # Convert mobile value to int if possible
        
        if phone_value_int in seen:
            combined_key = phone_value_int
        elif mobile_value_int in seen:
            combined_key = mobile_value_int
        else:
            combined_key = phone_value_int if phone_value_int is not None else mobile_value_int
        
        if combined_key is not None:
            if combined_key in duplicates_info:
                duplicates_info[combined_key].append((row[0].row, name_cell.value))
            else:
                duplicates_info[combined_key] = [(row[0].row, name_cell.value)]
            seen[combined_key] = row[0].row

    if duplicates_info:
        print("Lists of duplicates based on 'Phone' or 'Mobile':")
        for key, duplicates in duplicates_info.items():
            if len(duplicates) > 1:
                indices = [dup[0] for dup in duplicates]
                names = [dup[1] for dup in duplicates]
                print(f"Indices: {indices}, Names: {names}, Key: {key}")

                for index in sorted(indices[1:], reverse=True):
                    ws.delete_rows(index)
                    print(f"Deleted row {index}.")

    wb.save(file_path)
    print("Workbook saved. Duplicates removed.")

file_path = 'doctor_info.xlsx'
phone_column_name = 'Phone'
mobile_column_name = 'Mobile'

find_and_delete_duplicates(file_path, phone_column_name, mobile_column_name)
