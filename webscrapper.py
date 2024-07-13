import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import threading
import time
import psutil

def init_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("start-maximized")
    chrome_options.add_argument("disable-infobars")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("window-size=1920,1080")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

    return webdriver.Chrome(options=chrome_options)

def get_current_chrome_processes():
    chrome_pids = []
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'] in ('chrome.exe', 'chromedriver.exe'):
            chrome_pids.append(proc.info['pid'])
    return set(chrome_pids)

def kill_new_chrome_processes(initial_pids):
    current_pids = get_current_chrome_processes()
    new_pids = current_pids - initial_pids
    for pid in new_pids:
        try:
            proc = psutil.Process(pid)
            proc.kill()
        except psutil.NoSuchProcess:
            pass

def scrape_doctor_info(url):
    driver = init_driver()
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'CompanyNameLbl')))
        time.sleep(2)

        soup = BeautifulSoup(driver.page_source, 'html.parser')

        name_element = soup.find('label', id='CompanyNameLbl', class_='companyLabel_class', itemprop='name')
        name = name_element.span.text.strip() if name_element else None

        address_element = soup.find('label', id='AddressLbl')
        address = address_element.text.strip() if address_element else None

        profession_element = soup.find('label', id='ProfessionLbl', itemprop='description')
        profession = profession_element.text.strip() if profession_element else None

        phone_element = soup.find('label', class_='rc_firstphone')
        phone = phone_element.text.strip() if phone_element else None

        mobile_element = soup.find('label', id='MobileContLbl')
        mobile = mobile_element.span.text.strip() if mobile_element else None

        email_element = soup.find('a', rel='nofollow', class_='rc_Detaillink', href=True)
        email = email_element['href'].replace('mailto:', '') if email_element else None

        return name, address, profession, phone, mobile, email
    except Exception as e:
        print(f"Error loading page: {url}\n{e}")
        return None, None, None, None, None, None
    finally:
        driver.quit()
        time.sleep(2)

def scrape_and_append(url, df, lock):
    name, address, profession, phone, mobile, email = scrape_doctor_info(url)
    if name and not check_phone_exists(df, phone, mobile):
        with lock:
            doctor_data.append([name, address, profession, phone, mobile, email, ""])

def check_phone_exists(df, phone, mobile):
    phone = str(phone)
    mobile = str(mobile)
    return ((df["Phone"].astype(str) == phone) | (df["Mobile"].astype(str) == mobile)).any()

def get_doctor_links(driver, search_type, search_location):
    url = f"https://www.vrisko.gr/search/{search_type}/{search_location}"
    driver.get(url)

    soup = BeautifulSoup(driver.page_source, 'html.parser')
    doctor_links = []
    
    for div in soup.find_all('div', class_='AdvAreaRight'):
        website_link = div.find('a', class_='urlClickLoggingClass', target='_blank')
        if not website_link:
            more_link = div.find('a', class_='AdvAreaBottomRight')
            if more_link:
                doctor_links.append(more_link['href'])

    return doctor_links

def main():
    global doctor_data
    doctor_data = []
    lock = threading.Lock()

    initial_chrome_pids = get_current_chrome_processes()

    try:
        df = pd.read_excel('doctor_info.xlsx')
    except FileNotFoundError:
        df = pd.DataFrame(columns=['Name', 'Address', 'Profession', 'Phone', 'Mobile', 'Email', 'Ωra'])

    driver = init_driver()

    search_type = input("Enter the type of doctor (in Greek, e.g., Ορθοδοντικοί): ")
    search_location = input("Enter the location (in Greek, e.g., Αττική): ")

    doctor_links = get_doctor_links(driver, search_type, search_location)

    driver.quit()
    time.sleep(2)

    max_threads = 8
    active_threads = []

    for url in doctor_links:
        if not url:
            continue

        while len(active_threads) >= max_threads:
            for thread in active_threads:
                if not thread.is_alive():
                    active_threads.remove(thread)
            time.sleep(0.1)

        thread = threading.Thread(target=scrape_and_append, args=(url, df, lock))
        active_threads.append(thread)
        thread.start()

    for thread in active_threads:
        thread.join()

    updated_df = pd.DataFrame(doctor_data, columns=['Name', 'Address', 'Profession', 'Phone', 'Mobile', 'Email', 'Ωra'])

    df = pd.concat([df, updated_df]).drop_duplicates(subset=['Phone', 'Mobile'], keep='first')

    # Load the existing workbook and select the active sheet
    workbook = load_workbook('doctor_info.xlsx')
    sheet = workbook.active

    # Find the first empty row in the sheet
    row_start = sheet.max_row + 1

    for row in dataframe_to_rows(df, index=False, header=False):
        sheet.append(row)

    workbook.save('doctor_info.xlsx')
    print("Data appended to doctor_info.xlsx")

    kill_new_chrome_processes(initial_chrome_pids)

if __name__ == '__main__':
    main()
